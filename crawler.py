# coding=utf-8
# python 2 

import sys
import time
from BeautifulSoup import BeautifulSoup
import datetime
import re
import openpyxl
import requests
import smtplib

import email
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import os

def debugInputInfo() :
    print 'blackList'
    print blackList
    print 'whiteList'
    print whiteList
    print 'AllCVEList'
    print AllCVEList

def getHttp(url) :
    time.sleep(sTime)
    s = requests.Session()
    req = requests.Request('GET', url)
    prepped = req.prepare()
    prepped.headers['User-Agent'] = 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.99 Safari/537.36'
    count = 0
    while 1:
        if (count >= 3) :
            return response
        print 'Get : %s' % url
        response = s.send(prepped, verify=False)
        if (response.status_code == 200) :
            return response
        print response.status_code
        time.sleep(sTime)
        count += 1

def checkDate(begin, date, end) :
    if (begin <= date and date <= end) :
        return True
    return False

def inBlackList(title) :
    global blackList
    for black in blackList :
        if re.search(black, title, re.IGNORECASE) :
            return True
    return False

def checkCVE(CVE) :
    global AllCVEList
    if re.match('CVE-\d{4}-\d{4,7}', CVE) is None :
        return "NoCVE"
    if CVE in AllCVEList :
        return "CVErepeat"
    AllCVEList.append(CVE)
    return "New"

def filterCVEs(CVElist) :
    new = []
    for CVE in CVElist :
        status = checkCVE(CVE)
        if (status == "NoCVE") :
            return status
        if (status == "New") :
            new.append(CVE)
    return new

def checkCVEs(CVElist) :
    if (CVElist == "NoCVE") :
        return "NoCVE"
    if (CVElist == []) :
        return "CVErepeat"
    return "New"

def inWhiteList(title) :
    global whiteList
    for white in whiteList :
        if re.search(white, title, re.IGNORECASE) :
            return white
    return False

def getRisk(cve) :
    r = getHttp("https://web.nvd.nist.gov/view/vuln/detail?vulnId=" + cve)
    contents = BeautifulSoup(r.content).find("div", {
        "id": "p_lt_WebPartZone1_zoneCenter_pageplaceholder_p_lt_WebPartZone1_zoneCenter_VulnerabilityDetail_VulnFormView_Vuln3CvssPanel"
        })
    if re.search("vuln-cvssv3-base-score", str(contents)) == None :
        return "Not Found"
    firstRow = contents.findAll("a")[0].getText()
    if firstRow is not None :
        return firstRow.encode('utf-8')

def getRiskByCVElist(CVElist) :
    if (len(CVElist) == 1) :
        return getRisk(CVElist[0])
    riskMap = {"LOW" : 0, "MEDIUM" : 1, "HIGH" : 2}
    riskFlags = [0, 0, 0]
    for CVE in CVElist :
        risk = getRisk(CVE)
        if (risk == "Not Found") :
            continue
        riskFlags[riskMap[risk]] = 1
        if (riskFlags[0] == 1 and riskFlags[2] == 1) :
            break
    riskMap = dict((value, key) for key, value in riskMap.iteritems())
    risksLen = riskFlags.count(1)
    if risksLen == 0 :
        return "Not Found"
    elif risksLen == 1 :
        return riskMap[riskFlags.index(1)]
    elif risksLen == 2:
        output = ""
        first = 1
        for index, riskFlag in enumerate(riskFlags) :
            if (riskFlag == 1) :
                if (first == 1) :
                    output = riskMap[index] + " ~ "
                    first = 0
                else :
                    output = output + riskMap[index]
        return output
    elif risksLen == 3 :
        return riskMap[0] + " ~ " + riskMap[2]
    return "ERROR"

def riskEn2Tw(risk) :
    risk = risk.replace("LOW", "低")
    risk = risk.replace("MEDIUM", "中")
    risk = risk.replace("HIGH", "高")
    return risk

def setData(data) :
    global line
    for i in range(1, 8) :
        vuls.cell(row = line, column = i).value = data[i - 1]
    if (data[6] == "Black" or data[6] == "CVErepeat") :
        for i in range(1, 8) :
            vuls.cell(row = line, column = i).fill = grayFill
    elif (data[6] == "White" and data[4] == "") :
        for i in range(1, 8) :
            vuls.cell(row = line, column = i).fill = orangeFill
    elif (data[6] != "White" and line != 1) :
        for i in range(1, 8) :
            vuls.cell(row = line, column = i).fill = yellowFill
    line += 1

def getExploitDB(url) :
    global status_not_200
    r = getHttp(url)
    if (r.status_code <> 200 and status_not_200 < http_try_limit):
        status_not_200 = status_not_200 + 1
        return begin
    elif (status_not_200 == http_try_limit):
        return begin - datetime.timedelta(1)
    soup = BeautifulSoup(r.content)
    rows = soup.find("table").find("tbody").findAll("tr")
    date = 0
    for row in rows :
        cells = row.findAll("td")
        date = datetime.datetime.strptime(cells[0].getText(), "%Y-%m-%d")
        title = cells[4].getText()
        platform = cells[5].getText()
        source = cells[4].find('a').get('href')
        #if (source == "https://www.exploit-db.com/exploits/40988/") : continue #in case needed
        if (checkDate(begin, date, end) == False) :
            continue
        data = [date, title, platform, source, "", "", ""]
        if inBlackList(title) :
            data[6] = "Black"
            setData(data)
            continue
        sourceRequest = getHttp(source)
        sourceHttp = BeautifulSoup(sourceRequest.content)
        tdList = sourceHttp.find("table", {"class" : "exploit_list"}).findAll("td")
        aList = tdList[2].findAll("a")
        CVElist = "NoCVE"
        if (len(aList) != 0) :
            CVElistOrigin = re.findall('CVE-\d{4}-\d{4,7}', str(aList[0]))
            CVElist = filterCVEs(CVElistOrigin)
        data[6] = checkCVEs(CVElist)
        if (data[6] == "CVErepeat") :
            data[4] = ",".join(CVElistOrigin)
            setData(data)
            continue
        if (data[6] == "New") :
            data[4] = ",".join(CVElist)
            data[5] = riskEn2Tw(getRiskByCVElist(CVElist))
        platform = inWhiteList(title)
        if (platform != False) :
            data[2] = platform
            data[6] = "White"
        setData(data)
    return date

def getHkcert(url) :
    global status_not_200
    r = getHttp(url)
    if (r.status_code <> 200 and status_not_200 < http_try_limit):
        status_not_200 = status_not_200 + 1
        return begin
    elif (status_not_200 == http_try_limit):
        return begin - datetime.timedelta(1)
    soup = BeautifulSoup(r.content)
    rows = soup.find("table", attrs={"class": "sdchk_table3"}).find("tbody").findAll("tr")
    date = 0
    for row in rows :
        cells = row.findAll("td")
        date = datetime.datetime.strptime(cells[3].getText(), "%Y / %m / %d")
        title = cells[1].contents[0].getText()
        source = 'https://www.hkcert.org/' + str(cells[1].find('a').get('href'))
        if (checkDate(begin, date, end) == False) :
            continue
        data = [date, title, "", source, "", "", ""]
        if inBlackList(title) :
            data[6] = "Black"
            setData(data)
            continue
        sourceRequest = getHttp(source)
        sourceHttp = BeautifulSoup(sourceRequest.content)
        content6 = sourceHttp.find("div", {"id" : "content6"})
        CVElist = "NoCVE"
        if (content6 != None) :
            liList = content6.findAll("li")
            CVElistOrigin = []
            for li in liList :
                CVElistOrigin.append(li.getText())
            CVElist = filterCVEs(CVElistOrigin)
        data[6] = checkCVEs(CVElist)
        if (data[6] == "CVErepeat") :
            data[4] = ",".join(CVElistOrigin)
        if (data[6] != "New") :
            setData(data)
            continue
        platform = inWhiteList(title)
        if (platform != False) :
            data[2] = platform
            data[6] = 'White'
        data[4] = ",".join(CVElist)
        data[5] = riskEn2Tw(getRiskByCVElist(CVElist))
        setData(data)
    return date

def getNsfocus(url) :
    global status_not_200
    r = getHttp(url)
    if (r.status_code <> 200 and status_not_200 < http_try_limit):
        status_not_200 = status_not_200 + 1
        return begin
    elif (status_not_200 == http_try_limit):
        return begin - datetime.timedelta(1)
    r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text)
    rows = soup.find("ul", attrs={"class": "vul_list"}).findAll("li")
    date = 0
    for row in rows :
        # cn word print ERROR but save file OK
        date = datetime.datetime.strptime(row.find("span").getText(), "%Y-%m-%d")
        # save utf8 tw use excel import OK
        title = row.find("a").getText()
        source = "http://www.nsfocus.net" + str(row.find("a").get("href"))
        CVEnumber = ""
        CVEre = re.search('CVE-\d{4}-\d{4,7}', title)
        if (checkDate(begin, date, end) == False) :
            continue
        if (CVEre == None) :
            sourceRequest = getHttp(source)
            sourceHttp = BeautifulSoup(sourceRequest.content)
            CVEre = re.search('CVE-\d{4}-\d{4,7}', str(sourceHttp))
        if (CVEre != None) :
            CVEnumber = CVEre.group(0)
        title = title.replace("(" + CVEnumber + ")", "")
        data = [date, title, "", source, CVEnumber, "", ""]
        if inBlackList(title) :
            data[6] = "Black"
            setData(data)
            continue
        data[6] = checkCVE(CVEnumber)
        if (data[6] == "CVErepeat") :
            data[4] = CVEnumber
            setData(data)
            continue
        if (data[6] == "New") :
            data[4] = CVEnumber
            data[5] = riskEn2Tw(getRisk(CVEnumber))
        platform = inWhiteList(title)
        if (platform != False) :
            data[2] = platform
            data[6] = 'White'
        setData(data)
    return date

def crawlExploitDB() :
    urlList = [
        'https://www.exploit-db.com/remote/?order_by=date&order=desc',
        'https://www.exploit-db.com/webapps/?order_by=date&order=desc',
        'https://www.exploit-db.com/local/?order_by=date&order=desc',
        'https://www.exploit-db.com/dos/?order_by=date&order=desc'
    ]
    for url in urlList :
        pg = 1
        global status_not_200
        status_not_200 = 0
        while(1) :
            pgdate = getExploitDB(url + "&pg=" + str(pg))
            if (pgdate >= begin ) :
                pg += 1
            else :
                break

def crawlHkcert() :
    global status_not_200
    status_not_200 = 0
    hkcertURL = 'https://www.hkcert.org/security-bulletin?p_p_id=3tech_list_security_bulletin_full_WAR_3tech_list_security_bulletin_fullportlet&_3tech_list_security_bulletin_full_WAR_3tech_list_security_bulletin_fullportlet_cur='
    pg = 1
    while(1) :
        pgdate = getHkcert(hkcertURL + str(pg))
        if (pgdate >= begin) :
            pg += 1
        else :
            break

def crawlNsfocus() :
    global status_not_200
    status_not_200 = 0
    nsfocusURL = 'http://www.nsfocus.net/index.php?act=sec_bug'
    pg = 1
    while(1) :
        pgdate = getNsfocus(nsfocusURL + "&page=" + str(pg))
        if (pgdate >= begin) :
            pg += 1
        else :
            break

def send_mail(send_from, send_to, subject, text, files, server, port, username, password, isTls):
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = send_to
    msg['Date'] = email.utils.formatdate(localtime = True)
    msg['Subject'] = subject
    msg.attach(MIMEText(text))
    part = email.mime.base.MIMEBase('application', "octet-stream")
    part.set_payload(open(filename, "rb").read())
    email.encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename=%s' % (newFilename))
    msg.attach(part)
    smtp = smtplib.SMTP_SSL(server, port)
    if isTls:
        smtp.starttls()
    smtp.login(username,password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()

if __name__ == "__main__":
    #vvv [programming inside] vvv
    from requests.packages.urllib3.exceptions import InsecureRequestWarning
    requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
    from settings import settings ##########
    status_not_200 = 0
    data = ['Date', 'Title', 'Platform', 'Source', 'CVE', 'Risk', 'Status'] #資料 output 預設格式
    #^^^ [programming inside] ^^^
    
    print('\n---------- [ script start ]  {start_time} [ script start ] ----------'.format( start_time = datetime.datetime.now().strftime('%Y%m%d %H:%M:%S') ))
    sTime = settings.sTime
    path = os.path.abspath(settings.path) + '/'
    date_range_start = settings.date_range_start
    date_range_end = settings.date_range_end
    http_try_limit = settings.http_try_limit
    
    send_from = settings.send_from
    send_to = settings.send_to
    subject = settings.subject
    text = settings.text
    server = settings.server
    port = settings.port
    username = settings.username
    password = settings.password
    
    if (date_range_start > date_range_end ):
        print('plz setting date_range_start < date_range_end')
        exit
    today = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.min)
    begin = today - datetime.timedelta(days = date_range_start)
    end = today - datetime.timedelta(days = date_range_end)
    
    prevFilename = path + 'vulsList_' + begin.strftime('%Y%m%d') + '.xlsx'
    print('today : ' + today.strftime('%Y%m%d'))
    print('range : ' + begin.strftime('%Y%m%d') + ' ~ ' + end.strftime('%Y%m%d'))
    print('default except excel file : {}'.format(prevFilename))
    text = text.format(today = today.strftime('%Y%m%d') , begin = begin.strftime('%Y%m%d')+' 00:00:00 ' , end = end.strftime('%Y%m%d')+' 23:59:59 ' , crawl_time = datetime.datetime.now().strftime('%Y%m%d %H:%M:%S'))
    newFilename = 'vulsList_' + today.strftime('%Y%m%d') + '.xlsx'
    filename = path + newFilename
    try :
        wb = openpyxl.load_workbook(prevFilename)
    except IOError:
        if len(sys.argv) >1 :
            path = ''.join(sys.argv[1].split('\\')[:-2])
            wb = openpyxl.load_workbook(sys.argv[1])
        else:
            print('except a vaild excel file , e.g., {0}'.format(prevFilename))
            exit()
    oldSheet = wb.get_sheet_by_name('vulsHistory')
    wb.remove_sheet(oldSheet)
    vulsHistory = wb['vuls']
    vulsHistory.title = 'vulsHistory'
    vuls = wb.create_sheet(title='vuls')
    AllCVEList = []
    whiteList = []
    blackList = []
    nrows = vulsHistory.max_row + 1
    for i in range(2, nrows) :
        CVEs = str(vulsHistory.cell(row = i, column = 5).value)
        if "," in CVEs :
            AllCVEList = AllCVEList + CVEs.split(',')
        else :
            if (CVEs != "None") :
                AllCVEList.append(CVEs)
    wl = wb['whiteList']
    nrows = wl.max_row + 1
    for i in range(2, nrows) :
        whiteList.append(str(wl.cell(row = i, column = 1).value))
    bl = wb['blackList']
    nrows = bl.max_row + 1

    for i in range(2, nrows) :
        blackList.append(str(bl.cell(row = i, column = 1).value))
    
    grayFill = openpyxl.styles.PatternFill(start_color='FF969696', end_color='FF969696', fill_type='solid')
    orangeFill = openpyxl.styles.PatternFill(start_color='FFFFCC99', end_color='FFFFCC99', fill_type='solid')
    yellowFill = openpyxl.styles.PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
    line = 1
    #-------- call function --------#
    setData(data)
    crawlExploitDB()
    crawlHkcert()
    crawlNsfocus()
    runTime = today.strftime('%Y%m%d')
    newFilename = 'vulsList_' + runTime + '.xlsx'
    wb.save(path + newFilename)
    send_mail(send_from=send_from, send_to=send_to, subject=subject, text=text, files=filename, server=server, port=port, username=username, password=password, isTls=False)
    print('---------- [ script finish ]  {finish_time} [ script finish ] ----------\n'.format( finish_time = datetime.datetime.now().strftime('%Y%m%d %H:%M:%S') ))